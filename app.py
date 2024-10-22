from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify
import openpyxl
import pandas as pd
from sklearn.ensemble import RandomForestClassifier
from sklearn.preprocessing import OneHotEncoder
from sklearn.compose import ColumnTransformer
from sklearn.pipeline import Pipeline
from openpyxl import load_workbook
import random
import joblib
from predict_department import predict_department
from datetime import datetime, timedelta
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestClassifier
from sklearn.multioutput import MultiOutputClassifier
from constraint import Problem
import itertools
import os
app = Flask(__name__)
app.config['SECRET_KEY'] = 'your_secret_key_here'
PATIENT_EXCEL_FILE = 'DATA_BASE/PATIENTS_LOGIN_DETAILS.xlsx'
MANAGER_EXCEL_FILE = 'DATA_BASE/MANAGERS_LOGIN_DETAILS.xlsx'
STAFF_EXCEL_FILE = 'DATA_BASE/STAFF_LOGIN_DETAILS.xlsx'
SURGEONS_EXCEL_FILE = 'DATA_BASE/SURGEONS_LOGIN_DETAILS.xlsx'
# User model
class User:
    def __init__(self, username, password):
        self.username = username
        self.password = password
# Function to read users from Excel
def read_users_from_excel(excel_file):
    users = []
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            users.append(User(username=row[0], password=row[1]))
    except FileNotFoundError:
        pass
    return users
# Function to write users to Excel
def write_users_to_excel(users, excel_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for user in users:
        sheet.append([user.username, user.password])
    workbook.save(excel_file)
def initialize_manager_data():
    try:
        workbook = openpyxl.load_workbook(MANAGER_EXCEL_FILE)
        sheet = workbook.active
        if sheet.max_row <= 1:
            # If the sheet contains only headers, add a default manager
            default_manager = User(username='manager', password='password')
            write_users_to_excel([default_manager], MANAGER_EXCEL_FILE)
    except FileNotFoundError:
        # If the managers Excel file doesn't exist, create it and add a default manager
        default_manager = User(username='manager', password='password')
        write_users_to_excel([default_manager], MANAGER_EXCEL_FILE)
# Initialize manager data when the application starts
initialize_manager_data()
def read_availability_schedule(filename):
    workbook = load_workbook(filename)
    sheet = workbook.active
    availability_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        availability_data.append({
            'Name': row[0],
            'Department': row[2],
            'Slot Status': row[13],
            'Date': str(row[3]),
            'Time Slots': row[12]
        })
    return availability_data
# Patient login route
@app.route('/patient_login', methods=['GET', 'POST'])
def patient_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        patients = read_users_from_excel(PATIENT_EXCEL_FILE)
        for patient in patients:
            if patient.username == username and patient.password == password:
                session['username'] = username
                session['role'] = 'patient'
                return redirect(url_for('patient_dashboard'))
    return render_template('patient_login.html')
# Manager login route
@app.route('/manager_login', methods=['GET', 'POST'])
def manager_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        managers = read_users_from_excel(MANAGER_EXCEL_FILE)
        for manager in managers:
            if manager.username == username and manager.password == password:
                session['username'] = username
                session['role'] = 'manager'
                return redirect(url_for('manager_dashboard'))
        flash('Invalid username or password', 'error')
    return render_template('manager_login.html')
@app.route('/staff_login', methods=['GET', 'POST'])
def staff_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        role = request.form['role']
        if role == 'staff':
            users_excel_file = STAFF_EXCEL_FILE
        elif role == 'surgeon':
            users_excel_file = SURGEONS_EXCEL_FILE
        else:
            flash('Invalid role selected.', 'error')
            return redirect(url_for('staff_login'))
        users = read_users_from_excel(users_excel_file)
        for user in users:
            if user.username == username and user.password == password:
                session['username'] = username
                session['role'] = role
                if role == 'staff':
                    return redirect(url_for('staff_dashboard'))
                elif role == 'surgeon':
                    return redirect(url_for('surgeon_portal'))
        flash('Invalid username or password', 'error')
    return render_template('staff_login.html')
# Patient dashboard route
@app.route('/patient_dashboard')
def patient_dashboard():
    if 'username' in session and session['role'] == 'patient':
        return render_template('patient_dashboard.html', username=session['username'])
    else:
        flash('Please log in as a patient', 'error')
        return redirect(url_for('patient_login'))
# Manager dashboard route
@app.route('/manager_dashboard')
def manager_dashboard():
    if 'username' in session and session['role'] == 'manager':
        return render_template('manager_dashboard.html', username=session['username'])
    else:
        flash('Please log in as a manager', 'error')
        return redirect(url_for('manager_login'))
# Staff dashboard route
@app.route('/staff_dashboard')
def staff_dashboard():
    if 'username' in session and 'role' in session and session['role'] == 'staff':
        username = session['username']
        return render_template('staff_dashboard.html', user={'username': username})
    else:
        flash('Please log in as staff to access the dashboard', 'error')
        return redirect(url_for('staff_login'))
# Patient registration route
@app.route('/register_patient', methods=['GET', 'POST'])
def register_patient():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        patients = read_users_from_excel(PATIENT_EXCEL_FILE)
        for patient in patients:
            if patient.username == username:
                flash('Username already exists. Please choose a different username.', 'error')
                return redirect(url_for('register_patient'))
        new_patient = User(username=username, password=password)
        patients.append(new_patient)
        write_users_to_excel(patients, PATIENT_EXCEL_FILE)
        flash('Patient registered successfully. Please login.', 'success')
        return redirect(url_for('patient_login'))
    return render_template('register_patient.html')
# Staff registration route
@app.route('/register_staff', methods=['GET', 'POST'])
def register_staff():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        role = request.form['role']
        if role == 'staff':
            users_excel_file = STAFF_EXCEL_FILE
        elif role == 'surgeon':
            users_excel_file = SURGEONS_EXCEL_FILE
        else:
            return redirect(url_for('register_staff'))
        users = read_users_from_excel(users_excel_file)
        for user in users:
            if user.username == username:
                return redirect(url_for('register_staff'))
        new_user = User(username=username, password=password)
        users.append(new_user)
        write_users_to_excel(users, users_excel_file)
        return redirect(url_for('manager_dashboard'))
    return render_template('register_staff.html')
@app.route('/booking/')
def booking():
    availability_data = read_availability_schedule('DATA_BASE/DOCTOR_AVAILABILITY_DATA.xlsx')
    departments = set(item['Department'] for item in availability_data)
    return render_template('booking.html', departments=departments)
@app.route('/get_dates', methods=['POST'])
def get_dates():
    department = request.form['department']
    availability_data = read_availability_schedule('DATA_BASE/DOCTOR_AVAILABILITY_DATA.xlsx')
    dates = set(item['Date'] for item in availability_data if item['Department'] == department and item['Slot Status'] == 'Available')
    html = '<select id="date">'
    for date in dates:
        html += f'<option value="{date}">{date}</option>'
    html += '</select>'
    return html
@app.route('/get_times', methods=['POST'])
def get_times():
    date = request.form['date']
    availability_data = read_availability_schedule('DATA_BASE/DOCTOR_AVAILABILITY_DATA.xlsx')
    times = [item['Time Slots'] for item in availability_data if item['Date'] == date and item['Slot Status'] == 'Available']
    html = '<select id="time_slot">'
    for time_slot in times:
        html += f'<option value="{time_slot}">{time_slot}</option>'
    html += '</select>'
    return html
@app.route('/book_appointment', methods=['POST'])
def book_appointment():
    department = request.form.get('department')
    date = request.form.get('date')
    time_slot = request.form.get('time_slot')
    if not all([department, date, time_slot]):
        return "Error: Missing required fields"
    filename = 'DATA_BASE/DOCTOR_AVAILABILITY_DATA.xlsx'
    workbook = load_workbook(filename)
    sheet = workbook.active
    patient_id = random.randint(10000, 99999)
    booked = False
    doctor_name = None
    for row in range(2, sheet.max_row + 1):
        if (sheet.cell(row=row, column=3).value == department and 
            sheet.cell(row=row, column=4).value == date and
            sheet.cell(row=row, column=13).value == time_slot):
            sheet.cell(row=row, column=14).value = 'Booked'
            sheet.cell(row=row, column=8).value = patient_id 
            doctor_name = sheet.cell(row=row, column=1).value
            booked = True
            break
    workbook.save(filename)
    if booked:
    		message = f"Appointment booked successfully! Your patient ID is {patient_id}. Doctor: {doctor_name}."
    		return render_template("booking_result.html", success=True, message=message)
    else:
        message = "No available slots to book."
        return render_template("booking_result.html", success=False, message=message)
# Home page route
@app.route('/')
def home():
    return render_template('home.html')
@app.route('/predict_department_page')
def predict_department_page():
    return render_template('predict_department.html')
@app.route('/suggest', methods=['POST'])
def suggest():
    symptoms = request.form['symptoms']
    suggested_department = predict_department(symptoms)
    return render_template('department_result.html', department=suggested_department)
excel_path = 'DATA_BASE/DOCTOR_AVAILABILITY_DATA.xlsx'
def update_availability_doctor(df, index, new_status):
    try:
        if 0 <= index < len(df):
            df.at[index, 'Slot Status'] = new_status
            df.to_excel(excel_path, index=False)
            print(f"Updated index {index} to {new_status}")  
            return True
        else:
            print("Index out of range")  
            return False
    except Exception as e:
        print(f"Error updating the Excel: {e}")  
        return False

@app.route('/availability', methods=['GET', 'POST'])
def availability():
    if 'username' not in session or session.get('role') != 'staff':
        flash("Please log in to view schedules.")
        return redirect(url_for('staff_login'))
    doctor_name = session['username'] 
    df = pd.read_excel(excel_path)
    doctor_schedule = df[df['Name'].str.lower() == doctor_name.lower()].reset_index()
    schedule = doctor_schedule.to_dict('records')
    if request.method == 'POST':
        index_to_update = int(request.form['slot_index'])
        new_status = request.form['new_status']
        if update_availability_doctor(df, index_to_update, new_status):
            flash("Schedule updated successfully.")
            return redirect(url_for('availability'))
        else:
            flash("Error updating schedule. Please check the index.")
    return render_template('availability.html', schedule=schedule, doctor_name=doctor_name)
@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out successfully.', 'success')
    return redirect(url_for('staff_login'))
# Load the dataset
data_path = "DATA_BASE/Surgeon_Historical_Data.csv"
hospital_data = pd.read_csv(data_path)
departments = hospital_data['Department'].unique()
department_surgery_map = {department: hospital_data[hospital_data['Department'] == department]['Surgery Type'].unique().tolist() for department in departments}
@app.route('/surgery')
def surgery():
    return render_template('surgery.html', departments=departments, department_surgery_map=department_surgery_map)
X = hospital_data.drop(columns=["Surgeon Name", "Surgeon Rating"])
y = hospital_data["Surgeon Name"]
categorical_features = ["Gender", "Department", "Surgery Type"]
categorical_transformer = Pipeline(steps=[
    ('onehot', OneHotEncoder(handle_unknown='ignore'))
])
preprocessor = ColumnTransformer(
    transformers=[
        ('cat', categorical_transformer, categorical_features)
    ])
model = RandomForestClassifier(n_estimators=100, random_state=42)
pipeline = Pipeline(steps=[('preprocessor', preprocessor),
                           ('model', model)])
# Train the model
pipeline.fit(X, y)
# Function to predict surgeon based on input data
def predict_surgeon(patient_details):
    input_data = pd.DataFrame(patient_details, index=[0])
    surgeon = pipeline.predict(input_data)
    return surgeon[0]
@app.route('/predict', methods=['POST'])
def predict():
    if request.method == 'POST':
        # Collecting all patient details from the form
        patient_details = {
            "Patient Name": request.form['patient_name'],
            "Age": int(request.form['age']),
            "Gender": request.form['gender'],
            "Department": request.form['department'],
            "Surgery Type": request.form['surgery_type'],
            "Critical Level": int(request.form['critical_level']),
            "Diagnosis": request.form.get('diagnosis', '')  
        }
        predicted_surgeon = predict_surgeon(patient_details)
        surgeon_name = predicted_surgeon.split(' ')[-1]
        try:
            existing_data = pd.read_excel('DATA_BASE/SURGERY_SUGGESTIONS_DATA.xlsx')
        except FileNotFoundError:
            existing_data = pd.DataFrame(columns=[
                'Patient Name', 'Age', 'Gender', 'Department', 'Surgery Type', 'Critical Level', 'Diagnosis', 'Surgeon', 'Status'
            ])
        # Preparing new row for DataFrame
        new_row = pd.DataFrame({
            'Patient Name': [patient_details['Patient Name']],
            'Age': [patient_details['Age']],
            'Gender': [patient_details['Gender']],
            'Department': [patient_details['Department']],
            'Surgery Type': [patient_details['Surgery Type']],
            'Critical Level': [patient_details['Critical Level']],
            'Diagnosis': [patient_details['Diagnosis']],
            'Surgeon': [surgeon_name],
            'Status': ['Pending']
        })
        updated_data = pd.concat([existing_data, new_row], ignore_index=True)
        updated_data.to_excel('DATA_BASE/SURGERY_SUGGESTIONS_DATA.xlsx', index=False)
        return render_template('result.html', patient_name=patient_details['Patient Name'], predicted_surgeon=surgeon_name)
    return render_template('predict.html')
@app.route('/surgeon_portal')
def surgeon_portal():
    if 'username' in session:
        surgeon_name = session['username']
        surgeries_df = pd.read_excel('DATA_BASE/SURGERY_SUGGESTIONS_DATA.xlsx')
        surgeon_surgeries = surgeries_df[surgeries_df['Surgeon'] == surgeon_name]
        surgeries = surgeon_surgeries.to_dict(orient='records')
        return render_template('surgeon_portal.html', username=surgeon_name, surgeries=surgeries)
    else:
        flash('Please log in to access the surgeon portal.', 'error')
        return redirect(url_for('login'))
@app.route('/process_surgery', methods=['POST'])
def process_surgery():
    try:
        # Read the form data
        surgeon_name = request.form.get('surgeon_name')
        surgery_type = request.form.get('surgery_type')
        action = request.form.get('action')
        excel_file_path = 'DATA_BASE/SURGERY_SUGGESTIONS_DATA.xlsx'
        existing_data = pd.read_excel(excel_file_path)
        surgery_index = existing_data[
            (existing_data['Surgery Type'] == surgery_type) & 
            (existing_data['Surgeon'] == surgeon_name) &
            (existing_data['Status'] == 'Pending') 
        ].index[0]
        if action == 'accept':
            existing_data.loc[surgery_index, 'Status'] = 'Accepted'
            # Save the updated DataFrame back to the Excel file
            existing_data.to_excel(excel_file_path, index=False)
            # Redirect to the booking form after accepting the surgery
            return redirect(url_for('booking_form', surgeon_name=surgeon_name, surgery_type=surgery_type))
        elif action == 'reject':
            existing_data.loc[surgery_index, 'Status'] = 'Rejected'
        # Save the updated DataFrame back to the Excel file
        existing_data.to_excel(excel_file_path, index=False)
        # Redirect back to the surgeon portal after processing
        return redirect(url_for('surgeon_portal'))
    except Exception as e:
        # Log the error or display an error message
        print(f"An error occurred: {e}")
        return "An error occurred while processing the surgery.", 500  # Internal Server Error
@app.route('/booking_form/<surgeon_name>/<surgery_type>', methods=['GET', 'POST'])
def booking_form(surgeon_name, surgery_type):
    if request.method == 'POST':
        # Logic to process the submitted booking form
        return add_booking(request.form)
    else:
        # Render the form with pre-filled data
        return render_template('booking_form.html', surgeon_name=surgeon_name, surgery_type=surgery_type)
# Data loading functions
def load_data(filepath, usecols=None):
    return pd.read_csv(filepath, usecols=usecols)
# Prepare data for machine learning
def prepare_data():
    data = load_data("DATA_BASE/HISTORICAL_BOOKINGS_DATA.csv")
    staff_set = set()
    data['Assigned Supporting Staff'].dropna().str.split(',').apply(lambda x: staff_set.update([name.strip() for name in x]))
    for staff in staff_set:
        data[f'staff_{staff}'] = data['Assigned Supporting Staff'].str.contains(staff, regex=False).astype(int)
    feature_transformer = ColumnTransformer(
        [('surgery_type_ohe', OneHotEncoder(), ['Surgery Type'])],
        remainder='passthrough'
    )
    X = feature_transformer.fit_transform(data[['Surgery Type', 'Surgery Duration']])
    y = data[[f'staff_{staff}' for staff in staff_set]]
    return X, y, staff_set, feature_transformer
# Train the machine learning model
def train_model():
    X, y, staff_list, feature_transformer = prepare_data()
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.25, random_state=42)
    model = MultiOutputClassifier(RandomForestClassifier(n_estimators=100, random_state=42))
    model.fit(X_train, y_train)
    return model, feature_transformer, staff_list
# Global model and transformer
global_model, global_feature_transformer, global_staff_list = train_model()
def find_next_available_time(bookings, surgeon_name, proposed_date, surgeons):
    # Retrieve surgeon's working hours from the surgeon data
    surgeon_details = surgeons[surgeons['Surgeon Name'] == surgeon_name]
    if surgeon_details.empty:
        return None
    working_hours = surgeon_details.iloc[0]['Working Hours']
    shift_start_str, _ = working_hours.split('-')
    shift_start_time = datetime.strptime(shift_start_str.strip(), '%H:%M').time()
    latest_time = datetime.combine(proposed_date, shift_start_time)
    for _, booking in bookings.iterrows():
        booking_start = datetime.strptime(booking['Date and Time'], "%d-%m-%Y %H:%M")
        booking_end = booking_start + timedelta(hours=int(booking['Surgery Duration']))
        if booking['Assigned Surgeon'] == surgeon_name and booking_end > latest_time:
            latest_time = booking_end + timedelta(minutes=60)  
    return latest_time
def suggest_supporting_staff(surgery_type, num_staff_needed, model, feature_transformer, staff_list, staff_data, bookings, surgery_start_time, duration_hours):
    input_df = pd.DataFrame({'Surgery Type': [surgery_type], 'Surgery Duration': [duration_hours]})
    transformed_input = feature_transformer.transform(input_df)
    predictions = model.predict_proba(transformed_input)
    for attempt in range(3):  
        staff_probabilities = {staff: predictions[i][0][1] for i, staff in enumerate(staff_list)}
        available_staff = []
        for staff_name in sorted(staff_probabilities, key=staff_probabilities.get, reverse=True):
            if is_staff_available(bookings, staff_name, surgery_start_time, duration_hours):
                available_staff.append(staff_name)
                if len(available_staff) == num_staff_needed:
                    return available_staff
        input_df['Retry'] = attempt + 1
        transformed_input = feature_transformer.transform(input_df)
        predictions = model.predict_proba(transformed_input)
    return available_staff  
def is_staff_available(bookings, staff_name, surgery_start_time, duration_hours):
    proposed_end_time = surgery_start_time + timedelta(hours=duration_hours)
    for _, booking in bookings.iterrows():
        booking_start = datetime.strptime(booking['Date and Time'], "%d-%m-%Y %H:%M")
        booking_end = booking_start + timedelta(hours=int(booking['Surgery Duration']))
        booked_staff = booking['Assigned Supporting Staff'].split(', ')
        if staff_name in booked_staff and (booking_start < proposed_end_time and booking_end > surgery_start_time):
            return False
    return True
def parse_working_hours(hours_str):
    start_str, end_str = hours_str.split('-')
    start_time = datetime.strptime(start_str.strip(), '%H:%M').time()
    end_time = datetime.strptime(end_str.strip(), '%H:%M').time()
    return start_time, end_time
def is_time_available(bookings, start_time, end_time, surgeon_name, theatre_id, staff_members):
    for _, booking in bookings.iterrows():
        booking_start = datetime.strptime(booking['Date and Time'], "%d-%m-%Y %H:%M")
        booking_end = booking_start + timedelta(hours=int(booking['Surgery Duration']))
        if booking_start < end_time and booking_end > start_time and (
            booking['Assigned Surgeon'] == surgeon_name or
            booking['Theatre Assigned'] == theatre_id or
            any(staff in booking['Assigned Supporting Staff'].split(', ') for staff in staff_members)):
            return False
    return True
from flask import render_template
def add_booking(form_data):
    patient_name = form_data['patient_name']
    surgery_type = form_data['surgery_type']
    surgeon_name = form_data['surgeon_name']
    date = form_data['date']
    duration_hours = int(form_data['duration_hours'])
    num_support_staff = int(form_data['num_support_staff'])
    # Load booking, surgeon, and staff data
    bookings = load_data("DATA_BASE/NEW_BOOKINGS_DATA.csv")
    surgeons = load_data("DATA_BASE/SURGEON_AVAILABILITY_DATA.csv")
    staff_data = load_data("DATA_BASE/SUPPORTING_STAFF_AVAILABILITY_DATA.csv")
    # Find next available time for the surgeon
    surgeon_next_available_start = find_next_available_time(bookings, surgeon_name, datetime.strptime(date, "%d-%m-%Y"), surgeons)
    proposed_end_time = surgeon_next_available_start + timedelta(hours=duration_hours)
    # Suggest supporting staff using the ML model
    recommended_staff = suggest_supporting_staff(surgery_type, num_support_staff, global_model, global_feature_transformer, global_staff_list, staff_data, bookings, surgeon_next_available_start, duration_hours)
    if not recommended_staff:
        return jsonify({"message": "No available staff found for the proposed time."})
    # Prepare combinations of available staff
    staff_combinations = list(itertools.combinations(recommended_staff, num_support_staff))
    # Set up the constraint problem for scheduling
    problem = Problem()
    theatres = [{'Theatre ID': i} for i in range(101, 106)]  # Simplified theatre list
    problem.addVariable('theatre_id', [t['Theatre ID'] for t in theatres])
    problem.addVariable('staff_combination', staff_combinations)
    def constraint_function(theatre_id, staff_combination):
        return is_time_available(bookings, surgeon_next_available_start, proposed_end_time, surgeon_name, theatre_id, staff_combination)
    problem.addConstraint(constraint_function, ['theatre_id', 'staff_combination'])
    # Solve the problem to find valid booking options
    solutions = problem.getSolutions()
    if not solutions:
        return jsonify({"message": "No available slots found."})
    # Save the new booking if a valid solution is found
    solution = solutions[0]
    new_booking = {
        'Booking ID': len(bookings) + 1,
        'Date and Time': surgeon_next_available_start.strftime("%d-%m-%Y %H:%M"),
        'Assigned Surgeon': surgeon_name,
        'Surgery Type': surgery_type,
        'Assigned Supporting Staff': ', '.join(solution['staff_combination']),
        'Theatre Assigned': solution['theatre_id'],
        'Surgery Duration': duration_hours,
        'Number of Supporting Staff': num_support_staff,
        'Surgical Outcome': 'Pending',
        'Patient Recovery Time': 'Pending'
    }
    bookings = pd.concat([bookings, pd.DataFrame([new_booking])], ignore_index=True)
    bookings.to_csv('DATA_BASE/NEW_BOOKINGS_DATA.csv', index=False)
    new_booking_details = {
        'Booking ID': new_booking['Booking ID'],
        'Date and Time': new_booking['Date and Time'],
        'Assigned Surgeon': new_booking['Assigned Surgeon'],
        'Surgery Type': new_booking['Surgery Type'],
        'Assigned Supporting Staff': new_booking['Assigned Supporting Staff'],
        'Theatre Assigned': new_booking['Theatre Assigned'],
        'Surgery Duration': new_booking['Surgery Duration'],
        'Number of Supporting Staff': new_booking['Number of Supporting Staff'],
        'Surgical Outcome': new_booking['Surgical Outcome'],
        'Patient Recovery Time': new_booking['Patient Recovery Time']
    }
    # Return the rendered template with the new_booking_details
    return render_template("booking_details.html", new_booking_details=new_booking_details)
EXCEL_FILE = 'DATA_BASE/NEW_BOOKINGS_DATA.csv'
def clean_and_split_staff_names(staff_names):
    # Remove unwanted characters
    cleaned_names = staff_names.replace("[", "").replace("]", "").replace("'", "")
    # Split the names on comma and strip any surrounding whitespace
    split_names = [name.strip() for name in cleaned_names.split(',')]
    return split_names
@app.route('/schedule')
def schedule():
    # Load the data
    data = pd.read_csv(EXCEL_FILE)
    # Organize data by supporting staff
    staff_schedules = {}
    for index, row in data.iterrows():
        staff_list = clean_and_split_staff_names(row['Assigned Supporting Staff'])
        for staff in staff_list:
            if staff not in staff_schedules:
                staff_schedules[staff] = []
            staff_schedules[staff].append({
                'Date and Time': row['Date and Time'],
                'Surgery Type': row['Surgery Type'],
                'Theatre Assigned': row['Theatre Assigned']
            })
    # Organize data by theatre
    theatre_schedules = {}
    for index, row in data.iterrows():
        theatre = row['Theatre Assigned']
        if theatre not in theatre_schedules:
            theatre_schedules[theatre] = []
        theatre_schedules[theatre].append({
            'Date and Time': row['Date and Time'],
            'Surgery Type': row['Surgery Type'],
            'Supporting Staff': clean_and_split_staff_names(row['Assigned Supporting Staff'])
        })
    # Pass the organized data to the template
    return render_template('schedule.html', staff_schedules=staff_schedules, theatre_schedules=theatre_schedules)
@app.route('/surgeon_schedule/<user_name>')
def staff_schedule(user_name):  
    if 'username' in session:
        bookings = load_data("DATA_BASE/NEW_BOOKINGS_DATA.csv")
        # Filter data for surgeon schedules
        surgeon_datas = bookings[bookings['Assigned Surgeon'].str.contains(user_name)]
        return render_template('surgeon_schedule.html', schedules=surgeon_datas.to_dict(orient='records'))
    else:
        flash('Please log in to access the surgeon schedule.', 'error')
        return redirect(url_for('login'))  
if __name__ == '__main__':
    app.run(debug=True)
import itertools
from sklearn.multioutput import MultiOutputClassifier
from constraint import Problem
def load_data(filepath, usecols=None):
    return pd.read_csv(filepath, usecols=usecols)
def parse_input_date(date_str):
    try:
        return datetime.strptime(date_str, "%d/%m/%Y")
    except ValueError:
        print("Invalid date format. Please enter the date in DD/MM/YYYY format.")
        return None
# Prepare data for machine learning
def prepare_data(data):
    staff_list = set()
    data['Assigned Supporting Staff'].dropna().str.split(',').apply(lambda x: staff_list.update(clean_staff_names(','.join(x))))
    for staff in staff_list:
        data[f'staff_{staff}'] = data['Assigned Supporting Staff'].str.contains(staff, regex=False).astype(int)
    feature_transformer = ColumnTransformer(
        [('surgery_type_ohe', OneHotEncoder(), ['Surgery Type'])],
        remainder='passthrough'
    )
    X = feature_transformer.fit_transform(data[['Surgery Type', 'Surgery Duration']])
    y = data[[f'staff_{staff}' for staff in staff_list]]
    return X, y, staff_list, feature_transformer
# Train the machine learning model
def train_model(X, y, feature_transformer):
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.25, random_state=42)
    model = MultiOutputClassifier(RandomForestClassifier(n_estimators=100, random_state=42))
    model.fit(X_train, y_train)
    return model, feature_transformer 
def clean_staff_names(staff_names):
    names = staff_names.replace("[", "").replace("]", "").split(',')
    cleaned_names = set(name.strip() for name in names)
    return list(cleaned_names) 
def suggest_supporting_staff(surgery_type, num_staff_needed, model, feature_transformer, staff_list, staff_data, bookings, surgery_start_time, duration_hours):
    input_df = pd.DataFrame({'Surgery Type': [surgery_type], 'Surgery Duration': [duration_hours]})
    transformed_input = feature_transformer.transform(input_df)
    predictions = model.predict_proba(transformed_input)
    staff_probabilities = {}
    for i, staff in enumerate(staff_list):
        if len(predictions[i]) > 1:
            staff_probabilities[staff] = predictions[i][1][1]
        else:
            staff_probabilities[staff] = predictions[i][0][1]
    sorted_staff = sorted(staff_probabilities, key=staff_probabilities.get, reverse=True)
    available_staff = []
    for staff_name in sorted_staff:
        if is_staff_available(bookings, staff_name, surgery_start_time, duration_hours) and staff_name not in available_staff:
            available_staff.append(staff_name)
            if len(available_staff) == num_staff_needed:
                return available_staff
    if len(available_staff) < num_staff_needed:
        return f"Not enough available staff. Only {len(available_staff)} out of {num_staff_needed} required staff members are available."
    return available_staff